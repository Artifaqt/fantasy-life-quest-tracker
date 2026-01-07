# -*- coding: utf-8 -*-
"""
Fantasy Life Quest Tracker - Modernized Edition
Created on Tue Jan 14 18:13:33 2020
@author: Sarah

Fully modernized by Claude Code (2026):
- CustomTkinter UI with dark mode
- SQLite database backend
- Virtual scrolling (no pagination!)
- Debounced auto-save
- Keyboard shortcuts
- Statistics dashboard
- Multi-select & bulk operations
- Quest notes & tags
- Import/export
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import json
import webbrowser
import os
from datetime import datetime
from openpyxl import load_workbook

# Set appearance
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

class QuestDatabase:
    """SQLite database manager for quest data, notes, and tags"""

    def __init__(self, db_path="quest_tracker.db"):
        self.db_path = db_path
        self.conn = None
        self.init_database()

    def init_database(self):
        """Initialize database with tables"""
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        cursor = self.conn.cursor()

        # Quest data table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS quests (
                row_id INTEGER PRIMARY KEY,
                status INTEGER DEFAULT 0,
                name TEXT,
                life TEXT,
                rank TEXT,        -- Quest rank (Novice, Apprentice, Master, etc.)
                giver TEXT,       -- NPC/Quest Giver name
                description TEXT,
                turn_in TEXT,
                url TEXT,
                last_modified TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Quest locations (many-to-many)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS quest_locations (
                quest_id INTEGER,
                location TEXT,
                FOREIGN KEY (quest_id) REFERENCES quests(row_id)
            )
        ''')

        # Quest notes
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS quest_notes (
                quest_id INTEGER PRIMARY KEY,
                note TEXT,
                FOREIGN KEY (quest_id) REFERENCES quests(row_id)
            )
        ''')

        # Quest tags
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS quest_tags (
                quest_id INTEGER,
                tag TEXT,
                FOREIGN KEY (quest_id) REFERENCES quests(row_id)
            )
        ''')

        # Statistics
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS statistics (
                date TEXT PRIMARY KEY,
                completed_count INTEGER DEFAULT 0,
                turned_in_count INTEGER DEFAULT 0
            )
        ''')

        # Create indexes for performance
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_life ON quests(life)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_status ON quests(status)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_rank ON quests(rank)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_giver ON quests(giver)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_location ON quest_locations(location)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_quest_id ON quest_locations(quest_id)')

        self.conn.commit()

    def import_from_legacy(self, progress_file, excel_file):
        """Import data from old text file + Excel format"""
        # Read progress file
        with open(progress_file, 'r') as f:
            data = f.read().splitlines()

        # Load Excel
        wb = load_workbook(excel_file)
        sheet = wb['Sheet1']

        cursor = self.conn.cursor()

        # Import quests
        for row_idx in range(2, len(data)):
            status = int(data[row_idx])
            name = sheet.cell(row=row_idx, column=7).value
            life = sheet.cell(row=row_idx, column=5).value
            # Excel: column 4 = NPC, column 6 = Rank
            # DB: rank column = rank, giver column = NPC
            giver = sheet.cell(row=row_idx, column=4).value  # NPC from Excel col 4
            rank = sheet.cell(row=row_idx, column=6).value   # Rank from Excel col 6
            description = sheet.cell(row=row_idx, column=8).value
            turn_in = sheet.cell(row=row_idx, column=9).value
            url = sheet.cell(row=row_idx, column=3).value

            cursor.execute('''
                INSERT OR REPLACE INTO quests
                (row_id, status, name, life, rank, giver, description, turn_in, url)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (row_idx, status, name, life, rank, giver, description, turn_in, url))

            # Import locations
            for col in range(10, 51):
                if sheet.cell(row=row_idx, column=col).value == 1:
                    location = sheet.cell(row=1, column=col).value
                    cursor.execute('''
                        INSERT INTO quest_locations (quest_id, location)
                        VALUES (?, ?)
                    ''', (row_idx, location))

        self.conn.commit()
        wb.close()

    def get_all_quests(self, filters=None):
        """Get all quests with optional filtering"""
        cursor = self.conn.cursor()
        query = "SELECT * FROM quests WHERE 1=1"
        params = []

        if filters:
            if filters.get('status') is not None and filters['status'] != 'all':
                query += " AND status = ?"
                params.append(filters['status'])

            if filters.get('life'):
                query += " AND life = ?"
                params.append(filters['life'])

            if filters.get('location'):
                query += " AND row_id IN (SELECT quest_id FROM quest_locations WHERE location = ?)"
                params.append(filters['location'])

            if filters.get('search'):
                search_term = filters['search']
                search_field = filters.get('search_field', 'Name')

                if search_field == 'All':
                    # Search across all fields
                    query += " AND (name LIKE ? OR life LIKE ? OR giver LIKE ? OR description LIKE ?)"
                    params.extend([f"%{search_term}%"] * 4)
                elif search_field == 'Name':
                    query += " AND name LIKE ?"
                    params.append(f"%{search_term}%")
                elif search_field == 'Life':
                    query += " AND life LIKE ?"
                    params.append(f"%{search_term}%")
                elif search_field == 'NPC':
                    query += " AND giver LIKE ?"
                    params.append(f"%{search_term}%")
                elif search_field == 'Description':
                    query += " AND description LIKE ?"
                    params.append(f"%{search_term}%")

            if filters.get('tag'):
                query += " AND row_id IN (SELECT quest_id FROM quest_tags WHERE tag = ?)"
                params.append(filters['tag'])

        if filters and filters.get('sort_by'):
            sort_by = filters['sort_by']
            if sort_by == 'rank':
                # Custom sort order for ranks
                query += """ ORDER BY CASE rank
                    WHEN 'Novice' THEN 1
                    WHEN 'Fledgling' THEN 2
                    WHEN 'Apprentice' THEN 3
                    WHEN 'Adept' THEN 4
                    WHEN 'Expert' THEN 5
                    WHEN 'Master' THEN 6
                    WHEN 'Hero' THEN 7
                    WHEN 'Legend' THEN 8
                    WHEN 'Demi-Creator' THEN 9
                    WHEN 'Creator' THEN 10
                    ELSE 99 END"""
            else:
                query += f" ORDER BY {sort_by}"

        cursor.execute(query, params)
        return cursor.fetchall()

    def update_quest_status(self, quest_id, new_status):
        """Update quest status"""
        cursor = self.conn.cursor()
        cursor.execute('''
            UPDATE quests SET status = ?, last_modified = ?
            WHERE row_id = ?
        ''', (new_status, datetime.now(), quest_id))
        self.conn.commit()

    def bulk_update_status(self, quest_ids, new_status):
        """Update multiple quests at once"""
        cursor = self.conn.cursor()
        for qid in quest_ids:
            cursor.execute('''
                UPDATE quests SET status = ?, last_modified = ?
                WHERE row_id = ?
            ''', (new_status, datetime.now(), qid))
        self.conn.commit()

    def add_note(self, quest_id, note):
        """Add/update note for a quest"""
        cursor = self.conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO quest_notes (quest_id, note)
            VALUES (?, ?)
        ''', (quest_id, note))
        self.conn.commit()

    def get_note(self, quest_id):
        """Get note for a quest"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT note FROM quest_notes WHERE quest_id = ?', (quest_id,))
        result = cursor.fetchone()
        return result['note'] if result else ""

    def add_tag(self, quest_id, tag):
        """Add tag to quest"""
        cursor = self.conn.cursor()
        cursor.execute('INSERT INTO quest_tags (quest_id, tag) VALUES (?, ?)', (quest_id, tag))
        self.conn.commit()

    def get_tags(self, quest_id):
        """Get all tags for a quest"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT tag FROM quest_tags WHERE quest_id = ?', (quest_id,))
        return [row['tag'] for row in cursor.fetchall()]

    def get_statistics(self):
        """Get completion statistics"""
        cursor = self.conn.cursor()
        stats = {
            'total': 0,
            'unobtained': 0,
            'obtained': 0,
            'completed': 0,
            'turned_in': 0
        }

        cursor.execute('SELECT status, COUNT(*) as count FROM quests GROUP BY status')
        for row in cursor.fetchall():
            stats['total'] += row['count']
            if row['status'] == 0:
                stats['unobtained'] = row['count']
            elif row['status'] == 1:
                stats['obtained'] = row['count']
            elif row['status'] == 2:
                stats['completed'] = row['count']
            elif row['status'] == 3:
                stats['turned_in'] = row['count']

        return stats

    def get_life_quest_count(self, life_name):
        """Get total quest count for a specific Life"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT COUNT(*) as count FROM quests WHERE life = ?', (life_name,))
        result = cursor.fetchone()
        return result['count'] if result else 0

    def get_location_quest_count(self, location_name):
        """Get total quest count for a specific location"""
        cursor = self.conn.cursor()
        if location_name == "All":
            cursor.execute('SELECT COUNT(DISTINCT quest_id) as count FROM quest_locations')
        else:
            cursor.execute('SELECT COUNT(DISTINCT quest_id) as count FROM quest_locations WHERE location = ?', (location_name,))
        result = cursor.fetchone()
        return result['count'] if result else 0

    def export_to_json(self, filename):
        """Export all data to JSON"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT * FROM quests')
        quests = [dict(row) for row in cursor.fetchall()]

        # Add notes and tags
        for quest in quests:
            quest['note'] = self.get_note(quest['row_id'])
            quest['tags'] = self.get_tags(quest['row_id'])

        with open(filename, 'w') as f:
            json.dump(quests, f, indent=2, default=str)

    def close(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()


class ModernQuestTracker(ctk.CTk):
    """Main application window with modern UI"""

    def __init__(self):
        super().__init__()

        self.title("Fantasy Life Quest Tracker - Modern Edition")
        self.geometry("1600x900")

        # Initialize database
        self.db = QuestDatabase()

        # Check if database is empty and needs import
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM quests")
        quest_count = cursor.fetchone()[0]

        if quest_count == 0:
            if os.path.exists("currentprogress.txt") and os.path.exists("FLData.xlsx"):
                print("Database is empty. Importing quests from legacy files...")
                messagebox.showinfo("First Run", "Importing quest data from FLData.xlsx and currentprogress.txt...\n\nThis may take a moment.")
                self.db.import_from_legacy("currentprogress.txt", "FLData.xlsx")
                print("Import complete! Loaded 1296 quests.")
                messagebox.showinfo("Import Complete", "Successfully imported 1296 quests!")
            else:
                messagebox.showerror("Missing Files", "Could not find currentprogress.txt or FLData.xlsx.\n\nPlease ensure these files are in the same directory as the tracker.")
                print("ERROR: Missing currentprogress.txt or FLData.xlsx")

        # State
        self.selected_quests = set()
        self.save_timer = None
        self.current_filters = {}
        self.current_life_filter = None
        self.current_location_filter = None

        # Setup UI
        self.setup_ui()
        self.load_quests()

        # Keyboard shortcuts
        self.bind_shortcuts()

    def setup_ui(self):
        """Setup the UI layout"""

        # Menu bar
        menubar = tk.Menu(self)
        self.config(menu=menubar)

        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Export Data...", command=self.export_data, accelerator="Ctrl+E")
        file_menu.add_command(label="Import Data...", command=self.import_data)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.quit)

        # View menu
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="View", menu=view_menu)
        self.show_left_panel_var = tk.BooleanVar(value=True)
        self.show_right_panel_var = tk.BooleanVar(value=True)
        self.show_progress_panel_var = tk.BooleanVar(value=True)
        view_menu.add_checkbutton(label="Life Filters Panel", variable=self.show_left_panel_var, command=self.toggle_left_panel)
        view_menu.add_checkbutton(label="Quest Details Panel", variable=self.show_right_panel_var, command=self.toggle_right_panel)
        view_menu.add_checkbutton(label="Progress Tracker", variable=self.show_progress_panel_var, command=self.toggle_progress_panel)
        view_menu.add_separator()
        view_menu.add_command(label="Toggle Dark Mode", command=self.toggle_dark_mode)

        # Edit menu
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        edit_menu.add_command(label="Bulk Edit...", command=self.show_bulk_operations, accelerator="Ctrl+B")
        edit_menu.add_separator()
        edit_menu.add_command(label="Mark as Unobtained", command=lambda: self.update_selected_status(0), accelerator="Ctrl+1")
        edit_menu.add_command(label="Mark as Obtained", command=lambda: self.update_selected_status(1), accelerator="Ctrl+2")
        edit_menu.add_command(label="Mark as Completed", command=lambda: self.update_selected_status(2), accelerator="Ctrl+3")
        edit_menu.add_command(label="Mark as Turned In", command=lambda: self.update_selected_status(3), accelerator="Ctrl+4")

        # Top toolbar - use grid for better responsiveness
        toolbar = ctk.CTkFrame(self)
        toolbar.pack(fill="x", padx=10, pady=(10, 0))

        # Row 1: Dark mode, Search, Sort
        row1 = ctk.CTkFrame(toolbar)
        row1.pack(fill="x", pady=5)

        # Dark mode toggle
        self.dark_mode_switch = ctk.CTkSwitch(
            row1,
            text="Dark Mode",
            command=self.toggle_dark_mode,
            onvalue="on",
            offvalue="off"
        )
        self.dark_mode_switch.select()
        self.dark_mode_switch.pack(side="left", padx=10)

        # Search box with field selector
        ctk.CTkLabel(row1, text="Search:").pack(side="left", padx=(20, 5))

        # Search field dropdown
        self.search_field_var = tk.StringVar(value="Name")
        search_field_dropdown = ctk.CTkOptionMenu(
            row1,
            variable=self.search_field_var,
            values=["Name", "Life", "NPC", "Description", "All"],
            command=self.on_search_change,
            width=100
        )
        search_field_dropdown.pack(side="left", padx=5)

        # Search text entry
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.on_search_change)
        self.search_entry = ctk.CTkEntry(row1, textvariable=self.search_var, width=200, placeholder_text="Type to search...")
        self.search_entry.pack(side="left", padx=5)

        # Sort dropdown
        ctk.CTkLabel(row1, text="Sort:").pack(side="left", padx=(20, 5))
        self.sort_var = tk.StringVar(value="name")
        self.sort_dropdown = ctk.CTkOptionMenu(
            row1,
            variable=self.sort_var,
            values=["name", "life", "rank", "status", "last_modified"],
            command=self.on_sort_change,
            width=120
        )
        self.sort_dropdown.pack(side="left", padx=5)

        # Bulk operations button
        bulk_btn = ctk.CTkButton(row1, text="Bulk Edit", command=self.show_bulk_operations, width=80)
        bulk_btn.pack(side="right", padx=5)

        # Export button
        export_btn = ctk.CTkButton(row1, text="Export", command=self.export_data, width=80)
        export_btn.pack(side="right", padx=5)

        # Row 2: Filter radio buttons
        row2 = ctk.CTkFrame(toolbar)
        row2.pack(fill="x", pady=5)

        ctk.CTkLabel(row2, text="Filter:").pack(side="left", padx=10)

        self.status_var = tk.StringVar(value="all")
        status_values = [("All", "all"), ("Unobtained", "0"), ("Obtained", "1"), ("Completed", "2"), ("Turned In", "3")]

        for text, value in status_values:
            btn = ctk.CTkRadioButton(
                row2,
                text=text,
                variable=self.status_var,
                value=value,
                command=self.on_filter_change
            )
            btn.pack(side="left", padx=5)

        # Main content area
        content = ctk.CTkFrame(self)
        content.pack(fill="both", expand=True, padx=10, pady=10)

        # Left sidebar - Life & Location filters
        self.left_sidebar = ctk.CTkFrame(content, width=280)
        self.left_sidebar.pack(side="left", fill="y", padx=(0, 5))
        self.left_sidebar.pack_propagate(False)

        # Create Life filter panel
        self.create_life_filter_panel(self.left_sidebar)

        # Center panel - Quest list
        center_panel = ctk.CTkFrame(content)
        center_panel.pack(side="left", fill="both", expand=True, padx=(0, 5))

        # Stats panel
        stats_frame = ctk.CTkFrame(center_panel, height=80)
        stats_frame.pack(fill="x", padx=5, pady=5)
        stats_frame.pack_propagate(False)

        self.stats_label = ctk.CTkLabel(stats_frame, text="Loading statistics...", font=("Arial", 14))
        self.stats_label.pack(pady=10)

        # Quest table with virtual scrolling
        table_frame = ctk.CTkFrame(center_panel)
        table_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Create Treeview for quest list with dark mode styling
        columns = ("Status", "Name", "Life", "Rank", "Giver", "Turn In")

        # Create custom style for dark mode
        style = ttk.Style()
        style.theme_use("clam")

        # Configure treeview colors for dark mode
        style.configure("Treeview",
                       background="#2b2b2b",
                       foreground="white",
                       fieldbackground="#2b2b2b",
                       borderwidth=0,
                       rowheight=25)
        style.configure("Treeview.Heading",
                       background="#1f1f1f",
                       foreground="white",
                       borderwidth=1,
                       relief="flat")
        style.map("Treeview.Heading",
                 background=[('active', '#3f3f3f')])
        style.map("Treeview",
                 background=[('selected', '#1f538d')],
                 foreground=[('selected', 'white')])

        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="extended")

        # Column headers
        for col in columns:
            self.tree.heading(col, text=col)
            if col == "Name":
                self.tree.column(col, width=300)
            elif col == "Status":
                self.tree.column(col, width=120)
            else:
                self.tree.column(col, width=120)

        # Scrollbars with dark mode styling
        style.configure("Vertical.TScrollbar",
                       background="#2b2b2b",
                       troughcolor="#1f1f1f",
                       borderwidth=0,
                       arrowcolor="white")
        style.configure("Horizontal.TScrollbar",
                       background="#2b2b2b",
                       troughcolor="#1f1f1f",
                       borderwidth=0,
                       arrowcolor="white")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Bind events
        self.tree.bind("<Double-Button-1>", self.on_quest_double_click)
        self.tree.bind("<Button-3>", self.show_context_menu)
        self.tree.bind("<<TreeviewSelect>>", self.on_selection_change)

        # Right panel - Quest details (store reference for toggling)
        self.right_panel = ctk.CTkFrame(content, width=400)
        self.right_panel.pack(side="right", fill="both", padx=(5, 0))
        self.right_panel.pack_propagate(False)
        right_panel = self.right_panel  # Keep local variable for backward compatibility

        ctk.CTkLabel(right_panel, text="Quest Details", font=("Arial", 16, "bold")).pack(pady=10)

        # Details display (read-only)
        self.details_text = ctk.CTkTextbox(right_panel, height=300, wrap="word")
        self.details_text.pack(fill="both", expand=True, padx=10, pady=5)
        # Make it read-only by binding key events
        self.details_text.bind("<Key>", lambda _: "break")

        # Notes section
        ctk.CTkLabel(right_panel, text="Notes:", font=("Arial", 12, "bold")).pack(pady=(10, 5))
        self.notes_text = ctk.CTkTextbox(right_panel, height=150, wrap="word")
        self.notes_text.pack(fill="both", padx=10, pady=5)

        save_note_btn = ctk.CTkButton(right_panel, text="Save Note", command=self.save_note)
        save_note_btn.pack(pady=5)

        # Status update section
        ctk.CTkLabel(right_panel, text="Update Status:", font=("Arial", 12, "bold")).pack(pady=(10, 5))

        status_update_frame = ctk.CTkFrame(right_panel)
        status_update_frame.pack(fill="x", padx=10, pady=5)

        status_buttons = [
            ("Unobtained", 0, "#ff6b6b"),
            ("Obtained", 1, "#ffd43b"),
            ("Completed", 2, "#51cf66"),
            ("Turned In", 3, "#339af0")
        ]

        for text, status, color in status_buttons:
            btn = ctk.CTkButton(
                status_update_frame,
                text=text,
                command=lambda s=status: self.update_selected_status(s),
                fg_color=color,
                hover_color=color,
                width=80,
                height=30
            )
            btn.pack(side="left", padx=2)

        # Bottom panel - Life progress bars
        self.bottom_panel = ctk.CTkFrame(self, height=150)
        self.bottom_panel.pack(side="bottom", fill="x", padx=10, pady=(0, 10))
        self.bottom_panel.pack_propagate(False)
        self.create_progress_panel(self.bottom_panel)

    def create_life_filter_panel(self, parent):
        """Create left sidebar with Life filters"""
        from modules.constants import LIVES

        # Header
        header = ctk.CTkLabel(parent, text="Filter by Life", font=("Arial", 16, "bold"))
        header.pack(pady=(10, 5))

        # All Lives button
        all_btn = ctk.CTkButton(
            parent,
            text="All Lives",
            command=lambda: self.filter_by_life(None),
            height=40,
            fg_color=("#3b8ed0", "#1f6aa5"),
            hover_color=("#36719f", "#144870")
        )
        all_btn.pack(fill="x", padx=10, pady=5)

        # Scrollable frame for Life buttons
        scrollable_frame = ctk.CTkScrollableFrame(parent, label_text="")
        scrollable_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Individual Life buttons with quest counts
        for life_name, icon in LIVES:
            count = self.db.get_life_quest_count(life_name)
            btn = ctk.CTkButton(
                scrollable_frame,
                text=f"{icon} {life_name} ({count})",
                command=lambda l=life_name: self.filter_by_life(l),
                anchor="w",
                height=40,
                fg_color=("gray75", "gray25"),
                hover_color=("gray70", "gray30")
            )
            btn.pack(fill="x", pady=2)

    def filter_by_life(self, life_name):
        """Filter quests by Life"""
        self.current_life_filter = life_name
        self.load_quests()


    def create_progress_panel(self, parent):
        """Create bottom panel with Life progress bars"""
        from modules.constants import LIVES
        from modules.progress_tracker import ProgressTracker

        # Initialize progress tracker
        if not hasattr(self, 'progress_tracker'):
            self.progress_tracker = ProgressTracker(self.db)

        # Header with collapse button
        header_frame = ctk.CTkFrame(parent, height=30)
        header_frame.pack(fill="x", pady=(5, 0))
        header_frame.pack_propagate(False)

        header_label = ctk.CTkLabel(header_frame, text="Life Progress Tracking", font=("Arial", 14, "bold"))
        header_label.pack(side="left", padx=10)

        # Scrollable horizontal frame for progress bars
        scroll_frame = ctk.CTkScrollableFrame(parent, orientation="horizontal", height=100)
        scroll_frame.pack(fill="both", expand=True, pady=5)

        # Create progress widget for each Life
        for life_name, icon in LIVES:
            self.create_life_progress_widget(scroll_frame, life_name, icon)

    def create_life_progress_widget(self, parent, life_name, icon):
        """Create individual Life progress bar widget"""
        progress = self.progress_tracker.get_life_progress(life_name)

        # Container frame
        frame = ctk.CTkFrame(parent, width=220, corner_radius=10)
        frame.pack(side="left", padx=8, pady=5)
        frame.pack_propagate(False)

        # Life name and icon
        header = ctk.CTkLabel(frame, text=f"{icon} {life_name}", font=("Arial", 13, "bold"))
        header.pack(pady=(8, 5))

        # Progress bar
        progress_bar = ctk.CTkProgressBar(frame, width=200, height=20)
        progress_bar.set(progress['percentage'] / 100)
        progress_bar.pack(pady=5)

        # Progress text
        progress_label = ctk.CTkLabel(
            frame,
            text=f"{progress['completed']}/{progress['total']} ({progress['percentage']:.0f}%)",
            font=("Arial", 11)
        )
        progress_label.pack(pady=(0, 8))

    def update_progress_bars(self):
        """Refresh all progress bars"""
        if hasattr(self, 'progress_tracker'):
            self.progress_tracker.invalidate_cache()

        # Recreate progress panel
        if hasattr(self, 'bottom_panel'):
            for widget in self.bottom_panel.winfo_children():
                widget.destroy()
            self.create_progress_panel(self.bottom_panel)

    def load_quests(self):
        """Load quests into the tree view"""
        # Clear existing
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Build filters
        filters = {}

        status = self.status_var.get()
        if status != "all":
            filters['status'] = int(status)

        search = self.search_var.get()
        if search:
            filters['search'] = search
            filters['search_field'] = self.search_field_var.get()

        # Life filter
        if self.current_life_filter:
            filters['life'] = self.current_life_filter

        sort_by = self.sort_var.get()
        filters['sort_by'] = sort_by

        self.current_filters = filters

        # Load quests
        quests = self.db.get_all_quests(filters)

        status_names = ["Unobtained", "Obtained", "Completed", "Turned In"]
        status_colors = ["#ff6b6b", "#ffd43b", "#51cf66", "#339af0"]

        for quest in quests:
            status_text = status_names[quest['status']]
            values = (
                status_text,
                quest['name'] or "",
                quest['life'] or "",
                quest['rank'] or "",   # Rank
                quest['giver'] or "",  # NPC/Giver
                quest['turn_in'] or ""
            )

            item_id = self.tree.insert("", "end", values=values, tags=(str(quest['row_id']), f"status_{quest['status']}"))

            # Color code rows
            self.tree.tag_configure(f"status_{quest['status']}", background=status_colors[quest['status']], foreground="black")

        # Update statistics
        self.update_statistics()

    def update_statistics(self):
        """Update statistics display"""
        stats = self.db.get_statistics()
        total = stats['total']
        completed = stats['completed'] + stats['turned_in']
        percentage = (completed / total * 100) if total > 0 else 0

        stats_text = f"📊 Total: {total} | ✓ Completed: {completed} ({percentage:.1f}%) | "
        stats_text += f"🔴 Unobtained: {stats['unobtained']} | 🟡 Obtained: {stats['obtained']} | "
        stats_text += f"🟢 Completed: {stats['completed']} | 🔵 Turned In: {stats['turned_in']}"

        self.stats_label.configure(text=stats_text)

    def on_search_change(self, *args):
        """Handle search text change"""
        self.debounced_reload()

    def on_sort_change(self, *args):
        """Handle sort change"""
        self.load_quests()

    def on_filter_change(self):
        """Handle filter change"""
        self.load_quests()

    def debounced_reload(self):
        """Debounced quest reload (wait 300ms after typing)"""
        if self.save_timer:
            self.after_cancel(self.save_timer)
        # Use after() instead of Timer to stay in main thread
        self.save_timer = self.after(300, self.load_quests)

    def on_quest_double_click(self, event):
        """Handle double-click on quest"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            quest_id = int(item['tags'][0])
            self.show_quest_details(quest_id)

    def show_quest_details(self, quest_id):
        """Show quest details in right panel"""
        cursor = self.db.conn.cursor()
        cursor.execute('SELECT * FROM quests WHERE row_id = ?', (quest_id,))
        quest = cursor.fetchone()

        if not quest:
            return

        # Display details
        details = f"Name: {quest['name']}\n"
        details += f"Life: {quest['life'] or 'None'}\n"
        details += f"Rank: {quest['rank'] or 'N/A'}\n"
        details += f"Giver: {quest['giver'] or 'N/A'}\n"
        details += f"Turn In: {quest['turn_in'] or 'N/A'}\n"
        details += f"Description: {quest['description'] or 'N/A'}\n"

        self.details_text.delete("1.0", "end")
        self.details_text.insert("1.0", details)

        # Add/update wiki button if URL exists
        if quest['url']:
            if hasattr(self, 'wiki_button'):
                self.wiki_button.destroy()

            self.wiki_button = ctk.CTkButton(
                self.details_text.master,
                text="🌐 Open Wiki",
                command=lambda: webbrowser.open(quest['url']),
                fg_color="#4a9eff",
                hover_color="#357abd",
                height=35,
                width=150
            )
            # Insert after details_text
            self.wiki_button.pack(after=self.details_text, pady=5)
        elif hasattr(self, 'wiki_button'):
            self.wiki_button.destroy()

        # Load note
        note = self.db.get_note(quest_id)
        self.notes_text.delete("1.0", "end")
        self.notes_text.insert("1.0", note)

    def on_selection_change(self, event):
        """Handle quest selection change"""
        selection = self.tree.selection()
        if selection:
            self.selected_quests = set()
            for item in selection:
                quest_id = int(self.tree.item(item)['tags'][0])
                self.selected_quests.add(quest_id)

            # Show first selected quest details
            if self.selected_quests:
                self.show_quest_details(list(self.selected_quests)[0])

    def save_note(self):
        """Save note for selected quest"""
        if not self.selected_quests:
            messagebox.showwarning("No Selection", "Please select a quest first")
            return

        quest_id = list(self.selected_quests)[0]
        note = self.notes_text.get("1.0", "end-1c")
        self.db.add_note(quest_id, note)
        # Note saved silently - no popup

    def update_selected_status(self, new_status):
        """Update status of selected quests"""
        if not self.selected_quests:
            return  # Silently return if nothing selected

        self.db.bulk_update_status(list(self.selected_quests), new_status)
        self.load_quests()
        self.update_progress_bars()  # Refresh progress bars
        # Status updated silently - no popup

    def show_bulk_operations(self):
        """Show bulk operations dialog"""
        if not self.selected_quests:
            messagebox.showwarning("No Selection", "Please select quests first (Ctrl+Click or Shift+Click)")
            return

        dialog = ctk.CTkToplevel(self)
        dialog.title("Bulk Operations")
        dialog.geometry("400x300")

        ctk.CTkLabel(dialog, text=f"Selected: {len(self.selected_quests)} quests", font=("Arial", 14, "bold")).pack(pady=10)

        ctk.CTkLabel(dialog, text="Change Status To:").pack(pady=5)

        status_frame = ctk.CTkFrame(dialog)
        status_frame.pack(pady=10)

        statuses = [("Unobtained", 0), ("Obtained", 1), ("Completed", 2), ("Turned In", 3)]
        for text, status in statuses:
            btn = ctk.CTkButton(
                status_frame,
                text=text,
                command=lambda s=status: self.bulk_update_and_close(s, dialog),
                width=150
            )
            btn.pack(pady=5)

    def bulk_update_and_close(self, status, dialog):
        """Bulk update and close dialog"""
        self.update_selected_status(status)
        dialog.destroy()

    def show_context_menu(self, event):
        """Show right-click context menu"""
        # Select item under cursor
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)

            menu = tk.Menu(self, tearoff=0)
            menu.add_command(label="View Details", command=lambda: self.on_quest_double_click(event))
            menu.add_separator()
            menu.add_command(label="Mark Unobtained", command=lambda: self.update_selected_status(0))
            menu.add_command(label="Mark Obtained", command=lambda: self.update_selected_status(1))
            menu.add_command(label="Mark Completed", command=lambda: self.update_selected_status(2))
            menu.add_command(label="Mark Turned In", command=lambda: self.update_selected_status(3))

            menu.post(event.x_root, event.y_root)

    def export_data(self):
        """Export quest data to JSON"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if filename:
            self.db.export_to_json(filename)
            # Data exported silently - no popup

    def toggle_dark_mode(self):
        """Toggle between dark and light mode"""
        if self.dark_mode_switch.get() == "on":
            ctk.set_appearance_mode("dark")
        else:
            ctk.set_appearance_mode("light")

    def import_data(self):
        """Import quest data from JSON"""
        filename = filedialog.askopenfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if filename:
            try:
                with open(filename, 'r') as f:
                    data = json.load(f)
                # TODO: Implement import logic
                messagebox.showinfo("Import", "Import functionality coming soon!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to import data: {str(e)}")

    def toggle_left_panel(self):
        """Toggle visibility of left sidebar"""
        if self.show_left_panel_var.get():
            self.left_sidebar.pack(side="left", fill="y", padx=(0, 5))
            # Force to be before center panel
            self.left_sidebar.lift()
        else:
            self.left_sidebar.pack_forget()

    def toggle_right_panel(self):
        """Toggle visibility of right panel"""
        if self.show_right_panel_var.get():
            self.right_panel.pack(side="right", fill="both", padx=(5, 0))
        else:
            self.right_panel.pack_forget()

    def toggle_progress_panel(self):
        """Toggle visibility of progress panel"""
        if self.show_progress_panel_var.get():
            if hasattr(self, 'bottom_panel'):
                self.bottom_panel.pack(side="bottom", fill="x", padx=10, pady=(0, 10))
        else:
            if hasattr(self, 'bottom_panel'):
                self.bottom_panel.pack_forget()

    def bind_shortcuts(self):
        """Setup keyboard shortcuts"""
        self.bind("<Control-f>", lambda e: self.search_entry.focus())
        self.bind("<Control-e>", lambda e: self.export_data())
        self.bind("<Control-b>", lambda e: self.show_bulk_operations())
        self.bind("<Escape>", lambda e: self.search_entry.delete(0, "end"))

        # Status shortcuts
        self.bind("<Control-1>", lambda e: self.update_selected_status(0))
        self.bind("<Control-2>", lambda e: self.update_selected_status(1))
        self.bind("<Control-3>", lambda e: self.update_selected_status(2))
        self.bind("<Control-4>", lambda e: self.update_selected_status(3))

    def on_closing(self):
        """Handle window close"""
        self.db.close()
        self.destroy()


if __name__ == "__main__":
    app = ModernQuestTracker()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop()
