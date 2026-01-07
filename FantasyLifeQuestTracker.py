# -*- coding: utf-8 -*-
"""
Created on Tue Jan 14 18:13:33 2020

@author: Sarah

Modernized by Claude Code - Added sorting, search, caching, and performance improvements
"""

import tkinter as tk
from tkinter import *
from tkinter import ttk
from shutil import copyfile
import pandas as pd
from openpyxl import load_workbook
import sys
import webbrowser
import array
import io

global data #all 1296 quests here[0 and 1 are blank] currentprogress.txt
data = []
global placedata #all 48 [0,47] place names here placenames.txt
global topButton #which of the 5 btns [0. 4] at topright cur pressed
topButton = 4
global currentPlaceIndex #current place we are looking at, i.e. Lives
currentPlaceIndex = 47
global titleLocation #location for where to put the title Fantasy Life -
global wb #excel workbook FLData.xlsx
global text_scroll #for use to access text_scroll.frame to place data
global URLCol #column where the URl is displayed
URLCol = 3
global turnInCol #turn in location
turnInCol = 9
global livesCol #column that says the life
livesCol = 5
global livesIndex #index Lives is located at in placedata[]
livesIndex = 26
global allIndex #index all is located at in placedata[]
allIndex = 47
global startLocationIndex #where the location names start
startLocationIndex = 10
global endLocationIndex #where the location names end, PLUS 1 +1
endLocationIndex = 51
global dataIndexArray
global text_frame
global holdSelf
global nameCol #column the name occurs in
nameCol = 7
global choice_container
global minl
minl=0
global location_container_outer
global quest_cache  # Cached quest data from Excel to avoid repeated reads
quest_cache = []
global current_sort  # Current sort option: 'default', 'name', 'life', 'location', 'status'
current_sort = 'default'
global search_text  # Search filter text
search_text = ""
global sort_var  # StringVar for sort dropdown
global search_var  # StringVar for search box

def read():
    f = open("currentprogress.txt", "r")
    global data
    #data = f.readlines()
    data = f.read().splitlines() #removes the \n
    f.close()

def save():
    global data
    f = open("currentprogress.txt", "w")
    f.writelines("%s\n" % l for l in data)
    f.close()

def build_quest_cache():
    """Build a cache of all quest data from Excel to avoid repeated cell reads"""
    global quest_cache, wb, startLocationIndex, endLocationIndex, data
    quest_cache = []

    # Cache all quest data (rows 2 to len(data))
    for row_idx in range(2, len(data)):
        quest_info = {
            'row': row_idx,
            'status': int(data[row_idx]),
            'url': wb['Sheet1'].cell(row=row_idx, column=URLCol).value,
            'name': wb['Sheet1'].cell(row=row_idx, column=nameCol).value,
            'life': wb['Sheet1'].cell(row=row_idx, column=livesCol).value,
            'turn_in': wb['Sheet1'].cell(row=row_idx, column=turnInCol).value,
            'locations': []  # List of location names where quest is available
        }

        # Cache columns 2-9 for display
        quest_info['columns'] = {}
        for col in range(2, 10):
            quest_info['columns'][col] = wb['Sheet1'].cell(row=row_idx, column=col).value

        # Cache location availability
        for col in range(startLocationIndex, endLocationIndex):
            if wb['Sheet1'].cell(row=row_idx, column=col).value == 1:
                location_name = wb['Sheet1'].cell(row=1, column=col).value
                quest_info['locations'].append(location_name)

        quest_cache.append(quest_info)

def setText(obj):
        obj.text.set(str(obj.array[0]) + " / " +  str(obj.array[1]) + " / "  + str(obj.array[2]) + " / "  + str(obj.array[3]))
    
def findButtonName():
            global topButton
            if topButton == 4:
                return "All Requests"
            elif topButton == 3:
                return "Turned In Requests"
            elif topButton == 2:
                return "Completed Requests"
            elif topButton == 1:
                return "Obtained Requests"
            else:
                return "Unobtained Requests"
            
def changeTitle():
    global currentPlaceIndex
    global placedata
    titleLocation.title("Fantasy Life - {} - {}".format(placedata[currentPlaceIndex].key, findButtonName()))
  
def findLocation(name):
    global placedata
    for i in range(len(placedata)):
        if(placedata[i].key == name):
            return i
    return -1   

def findLocationCol(name):
    global startLocationIndex
    global endLocationIndex
    global wb
    for i in range(startLocationIndex, endLocationIndex):
        if(wb['Sheet1'].cell(row=1, column=i).value == name):
            return i
        
    return -1
    
def OpenUrl(i, *args):
    """Open the wiki URL for a quest - now uses cached data"""
    global quest_cache
    # Find quest in cache by row index
    for quest in quest_cache:
        if quest['row'] == i:
            if quest['url']:
                webbrowser.open_new(quest['url'])
            return

def locationcallback(i, *args):
    global location_container_outer
    global currentPlaceIndex
    global text_scroll
    global text_frame
    global holdSelf
    global topButton
    topButton = 1
    currentPlaceIndex = findLocation(location_container_outer[i-minl].get())
    changeTitle()
    text_frame.destroy()
    text_frame = Frame(holdSelf)
    text_frame.pack_propagate(0)
    text_frame.config(height=1000, width=1230)
    text_frame.pack()
    text_scroll = Scrollbar(text_frame, "Text")
    gatherData()
    

def callback(i, *args):
    """Handle quest status changes - now uses cached data for better performance"""
    global dataIndexArray, choice_container, placedata, quest_cache
    choices = ['Unobtained','Obtained','Completed','Turned In']

    row_idx = dataIndexArray[i]
    oldchoice = int(data[row_idx])
    newchoice = choices.index(choice_container[i-minl].get())
    data[row_idx] = str(newchoice)

    # Find quest in cache
    quest = None
    for q in quest_cache:
        if q['row'] == row_idx:
            quest = q
            quest['status'] = newchoice  # Update cache
            break

    if not quest:
        return

    locationName = quest['turn_in']
    curobj = placedata[findLocation(locationName)]

    if(oldchoice != 1): #unobtained, completed, turned in
        if (newchoice != 1):
            curobj.array[oldchoice] -= 1
            curobj.array[newchoice] += 1
            setText(curobj)
        else:
            curobj.array[oldchoice] -= 1
            # Use cached locations instead of Excel reads
            for loc_name in quest['locations']:
                tempobj = placedata[findLocation(loc_name)]
                tempobj.array[newchoice] += 1
                setText(tempobj)
    else:
        if(newchoice != 1):
            curobj.array[newchoice] += 1
            setText(curobj)
            # Use cached locations instead of Excel reads
            for loc_name in quest['locations']:
                tempobj = placedata[findLocation(loc_name)]
                tempobj.array[oldchoice] -= 1
                setText(tempobj)

    placedata[allIndex].array[oldchoice] -= 1
    placedata[allIndex].array[newchoice] += 1
    setText(placedata[allIndex])

    if quest['life'] is not None:
        placedata[livesIndex].array[oldchoice] -= 1
        placedata[livesIndex].array[newchoice] += 1
        setText(placedata[livesIndex])

    save()
    
def goBack(*args):
    global dataIndexArray
    global minl
    global text_frame
    global text_scroll
    if(minl - 29) >= 0:
        minl = minl - 29
        text_frame.destroy()
        text_frame = Frame(holdSelf)
        text_frame.pack_propagate(0)
        text_frame.config(height=1000, width=1230)
        text_frame.pack()
        text_scroll = Scrollbar(text_frame, "Text")
        showData()
    
def goForward(*args):
    global dataIndexArray
    global minl
    global text_frame
    global text_scroll
    tempmaxl = len(dataIndexArray)
    if(tempmaxl-minl) > 29:
        minl = minl + 29
        text_frame.destroy()
        text_frame = Frame(holdSelf)
        text_frame.pack_propagate(0)
        text_frame.config(height=1000, width=1230)
        text_frame.pack()
        text_scroll = Scrollbar(text_frame, "Text")
        showData()

def showData():
    """Display quest data - now uses cached data for better performance"""
    global text_scroll, dataIndexArray, choice_container, minl
    global location_container_outer, quest_cache, wb

    choice_container = []
    choices = ['Unobtained','Obtained','Completed','Turned In']
    location_container_outer = []

    ######Labels at the top of the info on right side######
    for j in range(2,10):
        if (j != URLCol):
            tk.Label(text_scroll.frame, text=wb['Sheet1'].cell(row=1, column=j).value).grid(row=0,column=j, sticky='nw')
    tk.Label(text_scroll.frame, text="Location").grid(row=0,column=10, sticky='nw')

    maxl = len(dataIndexArray)
    if (maxl-minl) > 29:
        maxl = minl+29

    for i in range(minl, maxl):
        row_idx = dataIndexArray[i]

        # Find quest in cache
        quest = None
        for q in quest_cache:
            if q['row'] == row_idx:
                quest = q
                break

        if not quest:
            continue

        # Status dropdown
        var = StringVar()
        choice_container.append(var)
        val = quest['status']
        choice_container[i-minl].set(choices[val])
        option = OptionMenu(text_scroll.frame, choice_container[i-minl], *choices)
        option.grid(row=i+1-minl, column=2)
        choice_container[i-minl].trace("w", lambda a, b, c, i=i: callback(i))

        # Display quest columns using cached data
        for j in range(3, 10):
            if (j != URLCol):
                if(j == nameCol):
                    obj=Button(text_scroll.frame, text=quest['name'])
                    obj.configure(command=lambda i=row_idx: OpenUrl(i))
                    obj.grid(row=i+1-minl,column=j, sticky='nw')
                else:
                    tk.Label(text_scroll.frame, text=quest['columns'][j]).grid(row=i+1-minl,column=j, sticky='nw')

        # Location dropdown - use cached locations
        if quest['locations']:
            locationvar = StringVar()
            locationvar.set(quest['locations'][0])
            locationoption = OptionMenu(text_scroll.frame, locationvar, *quest['locations'])
            locationoption.grid(row=i+1-minl, column=10)
            location_container_outer.append(locationvar)
            location_container_outer[i-minl].trace("w", lambda a, b, c, i=i: locationcallback(i))

    backb=Button(text_scroll.frame, text="<- Back")
    backb.configure(command=goBack)
    backb.grid(row=maxl+1-minl,column=4, sticky='nw')

    forwardb=Button(text_scroll.frame, text="Forward ->")
    forwardb.configure(command=goForward)
    forwardb.grid(row=maxl+1-minl,column=5, sticky='nw')

def gatherData():
    """Filter and sort quest data - now uses cached data with sorting support"""
    global currentPlaceIndex, topButton, placedata, allIndex, livesIndex
    global dataIndexArray, minl, quest_cache, current_sort, search_text

    dataIndexArray = []
    current_place_name = placedata[currentPlaceIndex].key

    # Filter quests based on status and location using cache
    for quest in quest_cache:
        row_idx = quest['row']
        status = quest['status']

        # Apply search filter
        if search_text and search_text.lower() not in quest['name'].lower():
            continue

        # Apply location and status filters
        include = False

        ##DISPLAYING ALL DATA includes "obtained"
        if(topButton == 4):
            if(currentPlaceIndex == allIndex): #if "All" is selected
                include = True
            elif((currentPlaceIndex == livesIndex) & (quest['life'] is not None)): #"Lives"
                include = True
            #If any location is selected
            elif((status != 1) & (current_place_name == quest['turn_in'])):
                include = True
            elif(status == 1) & (current_place_name in quest['locations']):
                include = True

        #UNOBTAINED, COMPLETED, TURNED IN
        elif((topButton != 1) & (status == topButton)):
            if(currentPlaceIndex == allIndex): #if "All" is selected
                include = True
            elif((currentPlaceIndex == livesIndex) & (quest['life'] is not None)): #"Lives"
                include = True
            elif(current_place_name == quest['turn_in']):
                include = True

        ##OBTAINED
        elif((topButton == 1) & (status == topButton)):
            if(currentPlaceIndex == allIndex): #if "All" is selected
                include = True
            elif((currentPlaceIndex == livesIndex) & (quest['life'] is not None)): #"Lives"
                include = True
            elif(current_place_name in quest['locations']):
                include = True

        if include:
            dataIndexArray.append(row_idx)

    # Apply sorting
    if current_sort == 'name':
        dataIndexArray.sort(key=lambda idx: next((str(q['name']) if q['name'] else '') for q in quest_cache if q['row'] == idx).lower())
    elif current_sort == 'life':
        dataIndexArray.sort(key=lambda idx: next((str(q['life']) if q['life'] else 'zzz') for q in quest_cache if q['row'] == idx).lower())
    elif current_sort == 'location':
        dataIndexArray.sort(key=lambda idx: next((str(q['turn_in']) if q['turn_in'] else '') for q in quest_cache if q['row'] == idx).lower())
    elif current_sort == 'status':
        dataIndexArray.sort(key=lambda idx: next(q['status'] for q in quest_cache if q['row'] == idx))
    # 'default' keeps original order

    minl = 0
    showData()
    

def initializeCount():
    """Initialize quest counters - now uses cached data for better performance"""
    global placedata, livesIndex, allIndex, quest_cache

    #0=unobtained 1=obtained 2=completed 3=turnedin
    for quest in quest_cache:
        status = quest['status']
        placedata[allIndex].array[status] += 1 #all array

        #if it is a Life quest
        if quest['life'] is not None:
            placedata[livesIndex].array[status] += 1

        #this covers adding the count to ALL sections except "obtained"
        if status != 1:
            locationIndex = findLocation(quest['turn_in'])
            if locationIndex >= 0:
                placedata[locationIndex].array[status] += 1
        else: #"obtained" update goes here
            for loc_name in quest['locations']:
                locationIndex = findLocation(loc_name)
                if locationIndex >= 0:
                    placedata[locationIndex].array[status] += 1

    #update the text for all locations
    for j in range(len(placedata)):
        setText(placedata[j])

def topB(v, *args):
    global topButton
    global text_scroll
    global text_frame
    global holdSelf
    topButton = v
    changeTitle()
    text_frame.destroy()
    text_frame = Frame(holdSelf)
    text_frame.pack_propagate(0)
    text_frame.config(height=1000, width=1230)
    text_frame.pack()
    text_scroll = Scrollbar(text_frame, "Text")
    gatherData()
    
def selectLocation(n, *args):
    global currentPlaceIndex
    global text_scroll
    global text_frame
    global holdSelf
    currentPlaceIndex = n
    changeTitle()
    text_frame.destroy()
    text_frame = Frame(holdSelf)
    text_frame.pack_propagate(0)
    text_frame.config(height=1000, width=1230)
    text_frame.pack()
    text_scroll = Scrollbar(text_frame, "Text")
    gatherData()

def onSortChange(*args):
    """Handle sort dropdown change"""
    global current_sort, sort_var, text_scroll, text_frame, holdSelf
    current_sort = sort_var.get().lower()
    text_frame.destroy()
    text_frame = Frame(holdSelf)
    text_frame.pack_propagate(0)
    text_frame.config(height=1000, width=1230)
    text_frame.pack()
    text_scroll = Scrollbar(text_frame, "Text")
    gatherData()

def onSearchChange(*args):
    """Handle search box change"""
    global search_text, search_var, text_scroll, text_frame, holdSelf
    search_text = search_var.get()
    text_frame.destroy()
    text_frame = Frame(holdSelf)
    text_frame.pack_propagate(0)
    text_frame.config(height=1000, width=1230)
    text_frame.pack()
    text_scroll = Scrollbar(text_frame, "Text")
    gatherData()

class ToolTip(object):

    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0

    def showtip(self, text):
        "Display text in tooltip window"
        self.text = text
        if self.tipwindow or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 27
        y = y + cy + self.widget.winfo_rooty() +27
        self.tipwindow = tw = Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry("+%d+%d" % (x, y))
        try:
            # For Mac OS
            tw.tk.call("::tk::unsupported::MacWindowStyle",
                       "style", tw._w,
                       "help", "noActivates")
        except TclError:
            pass
        label = Label(tw, text=self.text, justify=LEFT,
                      background="#ffffe0", relief=SOLID, borderwidth=1,
                      font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()
    
class Location():
    def __init__(self, widget):
        self.widget = widget
        self.photo = None
        self.label = None
        self.text = tk.StringVar()
        self.b = None
        self.key = None
        self.array = array.array('i',[0,0,0,0])  
    
    
    
class Scrollbar(tk.Frame):
    def scrollMap(self):
        ####This section is for initializing the left place buttons####
        fp = open("placenames.txt", "r")
        placenames = fp.read().splitlines()
        fp.close()
        
        fi = open("imagenames.txt", "r")
        imagenames = fi.read().splitlines()
        fi.close()
    
        x = 0
        y = 0
        global placedata
        placedata = [None] * 48
        for i in range(48):
            placedata[i] = Location(self)
            placedata[i].photo=PhotoImage(file=imagenames[i], master=self)
            placedata[i].label = Label(image=placedata[i].photo, master=self)
            placedata[i].label.image = placedata[i].photo # keep a reference!
            placedata[i].text = tk.StringVar()
            setText(placedata[i])
            placedata[i].b = Button(self.frame,textvariable=placedata[i].text,image=placedata[i].photo, command=lambda n=i: selectLocation(n), compound="top")
            placedata[i].b.grid(row=y, column=x)
            placedata[i].key = placenames[i]
            Window.createToolTip(placedata[i].b, placedata[i].key)
            
            x = x + 1
            if x == 6:
                x = 0
                y = y + 1
            
        ###########################################################
        
        initializeCount()
    
    def __init__(self, root, name_of_type):

        
        tk.Frame.__init__(self, root)
        self.canvas = tk.Canvas(root, borderwidth=0)
        self.frame = tk.Frame(self.canvas)
        self.vsb = tk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)
        if name_of_type != "Map":
            self.hsb = tk.Scrollbar(root, orient="horizontal", command=self.canvas.xview)
            self.canvas.configure(xscrollcommand=self.hsb.set)
            self.hsb.pack(side="bottom", fill="x")
            
        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.create_window((4,4), window=self.frame, anchor="nw", 
                                  tags="self.frame")
        
        self.frame.bind("<Configure>", self.onFrameConfigure) 
        if name_of_type == "Map":
            self.scrollMap()
    
    def onFrameConfigure(self, event):
        '''Reset the scroll region to encompass the inner frame'''
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        
    


class Window(Scrollbar):
    def __init__(self, master=None):
        Frame.__init__(self, master)               
        self.master = master
        self.init_window()
        
    def createToolTip(widget, text):
        toolTip = ToolTip(widget)
        def enter(event):
            toolTip.showtip(text)
        def leave(event):
            toolTip.hidetip()
        widget.bind('<Enter>', enter)
        widget.bind('<Leave>', leave)
        
    def init_window(self):
        # allowing the widget to take the full space of the root window
        self.pack(fill=BOTH, expand=1)
        global holdSelf
        holdSelf = self
        global titleLocation
        titleLocation = self.master

        map_frame = Frame(self)
        map_frame.pack_propagate(0)
        map_frame.config(height=1000, width=670)
        map_frame.pack(side = LEFT)

        map_scroll = Scrollbar(map_frame, "Map")

        button_frame = Frame(self)
        button_frame.pack()

        global text_frame
        text_frame = Frame(self)
        text_frame.pack_propagate(0)
        text_frame.config(height=1000, width=1230)
        text_frame.pack()

        global text_scroll
        text_scroll = Scrollbar(text_frame, "Text")
        changeTitle()
        gatherData()

        # creating filter buttons
        button_manager = [None]*5

        button_manager[0] = Button(button_frame, text="Unobtained Requests", command=lambda v=0: topB(v))
        button_manager[1] = Button(button_frame, text="Obtained Requests", command=lambda v=1: topB(v))
        button_manager[2] = Button(button_frame, text="Completed Requests", command=lambda v=2: topB(v))
        button_manager[3] = Button(button_frame, text="Turned In Requests", command=lambda v=3: topB(v))
        button_manager[4] = Button(button_frame, text="All Requests", command=lambda v=4: topB(v))

        # placing filter buttons
        for i in range(5):
            button_manager[i].grid(row = 0, column = i)

        # Add Sort dropdown
        global sort_var
        sort_var = StringVar()
        sort_var.set("Default")
        tk.Label(button_frame, text="Sort by:").grid(row=1, column=0, sticky='e', padx=(10,2))
        sort_options = ['Default', 'Name', 'Life', 'Location', 'Status']
        sort_dropdown = ttk.Combobox(button_frame, textvariable=sort_var, values=sort_options, state='readonly', width=15)
        sort_dropdown.grid(row=1, column=1, sticky='w', pady=5)
        sort_var.trace("w", onSortChange)

        # Add Search box
        global search_var
        search_var = StringVar()
        tk.Label(button_frame, text="Search:").grid(row=1, column=2, sticky='e', padx=(10,2))
        search_entry = tk.Entry(button_frame, textvariable=search_var, width=25)
        search_entry.grid(row=1, column=3, columnspan=2, sticky='w', pady=5)
        search_var.trace("w", onSearchChange)
            
        
        
    def client_exit(self):
        wb.close()
        exit() 

###############MAIN##################
        
try:
    f= open("currentprogress.txt")
except IOError:
    f= open("currentprogress.txt", "w+")
    f.write("-1\n-1\n")
    for i in range(1297):
        f.writelines("%s\n" %0)

f.close()
   
read()
save()


try:
    global wb
    wb = load_workbook('FLData.xlsx')
except FileNotFoundError:
    print("ERROR")
    sys.exit()

# Build quest cache for better performance
build_quest_cache()

root = Tk()

#size of the window
root.geometry("1500x1000")
root.iconbitmap(r'Images/icon.ico')


app = Window(root)

root.mainloop()
wb.close()